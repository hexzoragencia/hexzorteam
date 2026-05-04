#!/usr/bin/env python3
"""Escanea el Excel personal y muestra resumen de qué se va a importar."""
import openpyxl
import warnings
warnings.filterwarnings('ignore')
from datetime import datetime, date

XLSX = '/Users/macbookair/Downloads/_FINANZAS PERSONALES - ANUAL 2026 _ ORGANIZACION VICTOR ROSSO (1).xlsx'

MESES = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']
MES_NUM = {m: i+1 for i, m in enumerate(MESES)}

# Layout por mes (1-indexed cols)
LAYOUT = {
  'ingreso':         {'cols': {'nombre':2,  'fecha':3,  'esperado':4,  'real':5},  'rows': (21, 48)},
  'pago_programado': {'cols': {'nombre':9,  'fecha':10, 'esperado':11, 'real':12}, 'rows': (21, 32)},
  'suscripcion':     {'cols': {'nombre':9,  'fecha':10, 'esperado':11, 'real':12}, 'rows': (37, 48)},
  'gasto_mensual':   {'cols': {'nombre':15, 'fecha':None, 'esperado':16, 'real':17}, 'rows': (21, 48)},
  'ahorro':          {'cols': {'nombre':20, 'fecha':None, 'esperado':23, 'real':24}, 'rows': (21, 32)},
  'deuda':           {'cols': {'nombre':20, 'fecha':None, 'esperado':23, 'real':24}, 'rows': (37, 48)},
}
TX_ROWS_START = 53
TX_COLS = {'fecha':2, 'monto':3, 'rubro':5, 'descripcion':10}

def is_real_val(v):
    if v is None: return False
    if isinstance(v, str) and v.strip().upper() in ('TOTAL', 'NOMBRE GASTO', 'NOMBRE INGRESO', 'NOMBRE AHORRO', 'NOMBRE DEUDA', 'NOMBRE PAGO', 'SUSCRIPCIÓN'): return False
    if isinstance(v, bool): return False
    return True

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def to_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace(',','').replace('$',''))
        except: return 0
    return 0

wb = openpyxl.load_workbook(XLSX, data_only=True)

# 1. Recolectar categorías únicas por tipo
cats = {t: {} for t in LAYOUT}  # tipo -> { nombre: { presupuestos: {(año,mes): monto} } }
tx_count = {m: 0 for m in MESES}
tx_sample = []

for mes in MESES:
    if mes not in wb.sheetnames: continue
    ws = wb[mes]
    mn = MES_NUM[mes]

    for tipo, conf in LAYOUT.items():
        rstart, rend = conf['rows']
        cn = conf['cols']
        for r in range(rstart, rend+1):
            nombre = ws.cell(r, cn['nombre']).value
            esp = to_num(ws.cell(r, cn['esperado']).value) if cn['esperado'] else 0
            real = to_num(ws.cell(r, cn['real']).value) if cn['real'] else 0
            if not is_real_val(nombre): continue
            nombre = str(nombre).strip()
            if not nombre: continue
            cats[tipo].setdefault(nombre, {'presupuestos': {}})
            cats[tipo][nombre]['presupuestos'][(2026, mn)] = esp

    # Contar transacciones del Registro
    for r in range(TX_ROWS_START, ws.max_row+1):
        f = ws.cell(r, TX_COLS['fecha']).value
        m = ws.cell(r, TX_COLS['monto']).value
        if f and m:
            tx_count[mes] += 1
            if len(tx_sample) < 8:
                tx_sample.append({
                    'mes': mes,
                    'fecha': to_date(f),
                    'monto': to_num(m),
                    'rubro': ws.cell(r, TX_COLS['rubro']).value,
                    'desc': ws.cell(r, TX_COLS['descripcion']).value,
                })

print('='*70)
print('RESUMEN DE LO QUE SE VA A IMPORTAR')
print('='*70)
print()
print('CATEGORÍAS ÚNICAS POR TIPO:')
for tipo, items in cats.items():
    print(f'\n  [{tipo}] — {len(items)} únicas:')
    for nombre, info in sorted(items.items()):
        meses_con_pres = sum(1 for v in info['presupuestos'].values() if v > 0)
        total_anual = sum(info['presupuestos'].values())
        print(f'    • {nombre}  ({meses_con_pres} meses con presup, total año=${total_anual:,.0f})')

print()
print('TRANSACCIONES (Registro real) por mes:')
total_tx = 0
for m in MESES:
    if tx_count[m] > 0:
        print(f'  {m}: {tx_count[m]}')
        total_tx += tx_count[m]
print(f'  TOTAL: {total_tx}')

print()
print('Muestra de transacciones (primeras 8):')
for t in tx_sample:
    print(f'  {t["mes"]} {t["fecha"]} ${t["monto"]:,.0f} — {t["rubro"]} — {t["desc"] or ""}')

# 2. Fondos de reserva
print()
print('='*70)
print('FONDOS DE RESERVA')
print('='*70)
ws = wb['Fondos de reserva']
# Buscar la tabla
for r in range(1, 30):
    for c in range(1, 15):
        v = ws.cell(r, c).value
        if isinstance(v, str) and 'fondo' in v.lower() and len(v) < 40:
            print(f'  R{r} C{c}: {v!r}')
print()
print('Filas 1-20 col A-H:')
for r in range(1, 25):
    row = [str(ws.cell(r, c).value or '')[:25] for c in range(1, 9)]
    if any(row):
        print(f'  R{r}: ' + ' | '.join(row))

# 3. Deudas
print()
print('='*70)
print('CALCULADORA DE DEUDAS (snowball)')
print('='*70)
ws = wb['Calculadora de deudas - Bola de']
print('Filas 1-25 col A-H:')
for r in range(1, 25):
    row = [str(ws.cell(r, c).value or '')[:25] for c in range(1, 9)]
    if any(row):
        print(f'  R{r}: ' + ' | '.join(row))
