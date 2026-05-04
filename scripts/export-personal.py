#!/usr/bin/env python3
"""Lee el Excel personal y exporta data-personal.json para que import-personal.mjs lo cargue."""
import openpyxl, json, warnings
warnings.filterwarnings('ignore')
from datetime import datetime, date

XLSX = '/Users/macbookair/Downloads/_FINANZAS PERSONALES - ANUAL 2026 _ ORGANIZACION VICTOR ROSSO (1).xlsx'
OUT = '/Users/macbookair/Downloads/hexzor-empresarial/scripts/data-personal.json'

MESES = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']
MES_NUM = {m: i+1 for i, m in enumerate(MESES)}

LAYOUT = {
  'ingreso':         {'cols': {'nombre':2,  'fecha':3,  'esperado':4,  'real':5},  'rows': (21, 48)},
  'pago_programado': {'cols': {'nombre':9,  'fecha':10, 'esperado':11, 'real':12}, 'rows': (21, 32)},
  'suscripcion':     {'cols': {'nombre':9,  'fecha':10, 'esperado':11, 'real':12}, 'rows': (37, 48)},
  'gasto_mensual':   {'cols': {'nombre':15, 'fecha':None, 'esperado':16, 'real':17}, 'rows': (21, 48)},
  'ahorro':          {'cols': {'nombre':20, 'fecha':None, 'esperado':23, 'real':24}, 'rows': (21, 32)},
}
TX_ROWS_START = 54
TX_COLS = {'fecha':2, 'monto':3, 'rubro':5, 'descripcion':10}

# Excluir nombres que son ahorros de tipo "fondo de reserva" (van a fondos_reserva, no a categorías)
AHORROS_A_FONDO = {'CARRO', 'VIAJES'}
# Excluir como pago programado (es la cuota de una deuda — solo va como deuda)
EXCLUIR_PAGO_PROG = {'WESTER UNION'}

HEADERS = {'TOTAL', 'Nombre Gasto','Nombre ingreso','Nombre ahorro','Nombre deuda','Nombre pago','Suscripción','Subtotal','Subtotal:'}

def is_real(v):
    if v is None: return False
    if isinstance(v, str) and v.strip() in HEADERS: return False
    if isinstance(v, bool): return False
    return True

def to_iso_date(v):
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    return None

def to_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace(',','').replace('$',''))
        except: return 0
    return 0

wb = openpyxl.load_workbook(XLSX, data_only=True)

# 1. Recolectar categorías + presupuestos
cats_set = {}  # (tipo, nombre) -> orden
presupuestos = []

for mes in MESES:
    if mes not in wb.sheetnames: continue
    ws = wb[mes]
    mn = MES_NUM[mes]
    for tipo, conf in LAYOUT.items():
        rstart, rend = conf['rows']
        cn = conf['cols']
        for r in range(rstart, rend+1):
            nombre = ws.cell(r, cn['nombre']).value
            if not is_real(nombre): continue
            nombre = str(nombre).strip()
            if not nombre: continue
            # Excluir AHORROS que van a fondos
            if tipo == 'ahorro' and nombre.upper() in AHORROS_A_FONDO: continue
            # Excluir cuota WESTER UNION del pago_programado (solo va como deuda)
            if tipo == 'pago_programado' and nombre.upper() in EXCLUIR_PAGO_PROG: continue

            key = (tipo, nombre)
            if key not in cats_set:
                cats_set[key] = len(cats_set)

            esp = to_num(ws.cell(r, cn['esperado']).value) if cn['esperado'] else 0
            fecha = to_iso_date(ws.cell(r, cn['fecha']).value) if cn['fecha'] else None
            if esp > 0:
                presupuestos.append({
                    'tipo': tipo, 'nombre': nombre,
                    'ano': 2026, 'mes': mn,
                    'monto_esperado': esp,
                    'fecha_esperada': fecha
                })

categorias = [{'tipo': t, 'nombre': n, 'orden': o} for (t, n), o in cats_set.items()]

# 2. Transacciones
transacciones = []
for mes in MESES:
    if mes not in wb.sheetnames: continue
    ws = wb[mes]
    for r in range(TX_ROWS_START, ws.max_row+1):
        f = ws.cell(r, TX_COLS['fecha']).value
        m = ws.cell(r, TX_COLS['monto']).value
        if not (isinstance(f, datetime) and isinstance(m, (int, float)) and m > 0):
            continue
        rubro = ws.cell(r, TX_COLS['rubro']).value
        desc = ws.cell(r, TX_COLS['descripcion']).value
        transacciones.append({
            'fecha': to_iso_date(f),
            'monto': float(m),
            'rubro': str(rubro).strip() if rubro else None,
            'descripcion': str(desc).strip() if desc else None,
        })

# 3. Fondos de reserva: CARRO y VIAJES con meta $80M
fondos = [
    {'nombre': 'CARRO',  'meta': 80_000_000, 'pago_mensual': 0, 'inicial': 0, 'ahorrado': 0},
    {'nombre': 'VIAJES', 'meta': 80_000_000, 'pago_mensual': 0, 'inicial': 0, 'ahorrado': 0},
]

# 4. Deudas: leer la calculadora snowball
deudas = []
ws = wb['Calculadora de deudas - Bola de']
for r in range(9, 22):
    nombre = ws.cell(r, 2).value
    balance = ws.cell(r, 4).value
    interes = ws.cell(r, 5).value
    cuota = ws.cell(r, 6).value
    if isinstance(nombre, str) and nombre.strip().upper() not in ('TOTAL','NOMBRE DE LA DEUDA','') and isinstance(balance, (int, float)) and balance > 0:
        deudas.append({
            'nombre': nombre.strip(),
            'balance': float(balance),
            'interes_anual': float(interes) if isinstance(interes, (int, float)) else 0,
            'cuota_mensual': float(cuota) if isinstance(cuota, (int, float)) else 0,
        })

out = {
    'categorias': categorias,
    'presupuestos': presupuestos,
    'transacciones': transacciones,
    'fondos': fondos,
    'deudas': deudas,
}

with open(OUT, 'w') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f'✅ Exportado a {OUT}')
print(f'   Categorías: {len(categorias)}')
print(f'   Presupuestos: {len(presupuestos)}')
print(f'   Transacciones: {len(transacciones)}')
print(f'   Fondos: {len(fondos)}')
print(f'   Deudas: {len(deudas)}')
