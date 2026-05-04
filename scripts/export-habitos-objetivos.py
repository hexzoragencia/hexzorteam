#!/usr/bin/env python3
"""Lee los Excels de hábitos y objetivos y exporta data-habitos-objetivos.json."""
import openpyxl, json, warnings
warnings.filterwarnings('ignore')

XLSX_HABITOS = '/Users/macbookair/Downloads/HABITOS DE CRECIMIENTO 2026 (1).xlsx'
XLSX_OBJETIVOS = '/Users/macbookair/Downloads/MIS OBJETIVIOS 2026.xlsx'
OUT = '/Users/macbookair/Downloads/hexzor-empresarial/scripts/data-habitos-objetivos.json'

# ===== HÁBITOS: extraer la lista desde la hoja "Ene" (col B, filas 10+) =====
wb = openpyxl.load_workbook(XLSX_HABITOS, data_only=True)
ws = wb['Ene']
EXCLUIR = {'progreso','hecho','no hecho','total','ánimo','animo','motivación','motivacion'}
habitos = []
for r in range(10, 23):  # límite duro: hábitos están en R10-22 según el Excel
    nombre = ws.cell(r, 2).value
    if isinstance(nombre, str) and nombre.strip() and nombre.strip().lower() not in EXCLUIR:
        habitos.append({'nombre': nombre.strip()})

# ===== OBJETIVOS: leer "LINEA DE 5 MIL DOLARES" (proyectos financieros) =====
wb = openpyxl.load_workbook(XLSX_OBJETIVOS, data_only=True)
ws = wb['LINEA DE 5 MIL DOLARES']
objetivos = []
# Las filas tienen: A=numero, B=nombre proyecto, F=ingreso, G=cantidad, K=ganancia
for r in range(8, 50):
    num = ws.cell(r, 1).value
    nombre = ws.cell(r, 2).value
    if not (isinstance(num, (int, float)) and isinstance(nombre, str) and nombre.strip()):
        continue
    ingreso = ws.cell(r, 6).value
    cantidad = ws.cell(r, 7).value
    ganancia = ws.cell(r, 11).value
    obj = {
        'titulo': nombre.strip(),
        'tipo': 'financiero',
        'estado': 'en_progreso',
        'ingreso_esperado': float(ingreso) if isinstance(ingreso, (int, float)) else None,
        'ganancia_esperada': float(ganancia) if isinstance(ganancia, (int, float)) else None,
        'cantidad': int(cantidad) if isinstance(cantidad, (int, float)) and cantidad > 0 else 1,
        'progreso': 0,
    }
    objetivos.append(obj)

with open(OUT, 'w') as f:
    json.dump({'habitos': habitos, 'objetivos': objetivos}, f, indent=2, ensure_ascii=False)

print(f'✅ Exportado a {OUT}')
print(f'   Hábitos: {len(habitos)}')
for h in habitos: print(f'     • {h["nombre"]}')
print(f'   Objetivos: {len(objetivos)}')
for o in objetivos:
    extra = f' — ingreso ${o["ingreso_esperado"]:,.0f} × {o["cantidad"]}' if o["ingreso_esperado"] else ''
    print(f'     • {o["titulo"]}{extra}')
